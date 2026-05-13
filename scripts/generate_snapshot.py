import json
import os
import io
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import openpyxl
from datetime import datetime, date, timedelta

# Auth
creds_json = os.environ['GOOGLE_SERVICE_ACCOUNT_KEY']
creds_dict = json.loads(creds_json)
creds = Credentials.from_service_account_info(
    creds_dict,
    scopes=['https://www.googleapis.com/auth/drive.readonly']
)

drive = build('drive', 'v3', credentials=creds)
SHEET_ID = '16gycwzxACC2--gNuWpGeN0kcjtXUGv1d'

XLSX_MIME = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

def download_workbook(file_id):
    meta = drive.files().get(fileId=file_id, fields='mimeType').execute()
    if meta['mimeType'] == 'application/vnd.google-apps.spreadsheet':
        request = drive.files().export_media(fileId=file_id, mimeType=XLSX_MIME)
    else:
        request = drive.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return openpyxl.load_workbook(buf, data_only=True)

wb = download_workbook(SHEET_ID)

# Static enrollment + YTD pacing — baked once from the cohort workbook so the
# daily refresh only needs to read the BTS Tracker.
STATIC_METRICS_PATH = os.path.join(os.path.dirname(__file__), 'static-metrics.json')
try:
    with open(STATIC_METRICS_PATH) as f:
        static_metrics = json.load(f)
except FileNotFoundError:
    static_metrics = {}

def get_static_metrics(name):
    key = name.lower()
    if key in static_metrics:
        return static_metrics[key]
    for k, v in static_metrics.items():
        if k in key or key in k:
            return v
    return {'enrollment': None, 'ytdPacing': None}

# Static upsell flags — baked one-time from Looker (District Health Plan, pulled
# 2026-05-12). Not refreshed; upsell conversations happen once per district
# during EOY meetings, so flags must stay consistent through the season.
STATIC_UPSELL_PATH = os.path.join(os.path.dirname(__file__), 'static-upsell.json')
try:
    with open(STATIC_UPSELL_PATH) as f:
        static_upsell = json.load(f)
except FileNotFoundError:
    static_upsell = {}

def get_static_upsell(name):
    # Exact match only — substring matching was too permissive (e.g., "union
    # elementary school district" silently flagging "Alisal Union Elementary").
    # If a curated upsell name doesn't exactly match a tracker name, fix the
    # name in static-upsell.json (or the tracker) so the join is explicit.
    return static_upsell.get(name.strip().lower())

# Find District Tracker sheet. Other sheets (Monicas Sheet, Daisys Sheet, How
# to Use) live in their own tabs and are ignored here.
sheet = None
for name in wb.sheetnames:
    if 'district tracker' in name.lower():
        sheet = wb[name]
        break
if sheet is None:
    sheet = wb.active

rows = []
for row in sheet.iter_rows(values_only=True):
    rows.append([str(cell) if cell is not None else '' for cell in row])

today = date.today()

def parse_date(s):
    if not s or s.strip() in ['', '#VALUE!', 'N/A', 'not stated in website', 'None']:
        return None
    s = s.strip()
    for fmt in ['%m/%d/%Y', '%m/%d/%y', '%m-%d-%Y', '%m-%d-%y']:
        try:
            d = datetime.strptime(s, fmt)
            if d.year > 2027:
                return None
            return d.date()
        except:
            pass
    return None

def booking_target(ldos):
    return ldos - timedelta(days=28) if ldos else None

def yn(val):
    return str(val).strip().lower() in ['y', 'yes', 'true', '1'] if val else False

def is_overdue(booked, completed, bt):
    if completed or booked or not bt:
        return False
    return today > bt

def needs_nudge(booked, completed, outreach, last_outreach):
    if completed or booked or not outreach or not last_outreach:
        return False
    return (today - last_outreach).days >= 3

def get_status(completed, meeting_type, overdue, booked, nudge):
    if completed: return 'completed'
    if meeting_type == 'async': return 'async'
    if overdue: return 'overdue'
    if booked: return 'booked'
    if nudge: return 'nudge'
    return 'schedule-soon'

def parse_tier(val):
    s = str(val or '').strip().lower()
    if not s:
        return 3
    if '1' in s: return 1
    if '2' in s: return 2
    if '3' in s: return 3
    return 3

def parse_meeting_type(val, tier_num):
    if val:
        s = str(val).strip().lower()
        if s in ('y', 'yes', 'true', '1', 'live'):
            return 'live'
        if s in ('n', 'no', 'false', '0', 'async'):
            return 'async'
    return 'async' if tier_num == 3 else 'live'

# Find header row
header_idx = None
for i, row in enumerate(rows):
    if row and row[0] == 'District Name' and len(row) > 5:
        header_idx = i
        break

if header_idx is None:
    raise Exception("Could not find header row")

# Parse district rows. Section-header break logic was removed — sections that
# used to live in this tab now live in their own tabs (Monicas Sheet, etc.),
# and the substring break was misfiring on real district names like
# "Santa Monica-Malibu" and "Del Norte" (no Monica/Daisy in those, but the
# substring trigger caught any name that happened to contain them).
data_rows = []
for row in rows[header_idx + 1:]:
    if not row or not row[0].strip():
        continue
    data_rows.append(row)

csm_map = {
    'Brianna Masciel': 'brianna',
    'Sarah Hough': 'sarah',
    'Monica Knott': 'monica',
    'Daisy Leahy': 'daisy',
}

def empty_urgency_buckets():
    return {
        'thisWeek': [],
        'nextWeek': [],
        'soon': [],
        'later': [],
        'noDate': [],
    }

def empty_tier_stats():
    return {'total': 0, 'completed': 0, 'booked': 0, 'remaining': 0, 'overdue': 0}

def empty_gap_to_goal():
    return {
        'totalNeedingCall': 0,
        'booked': 0,
        'completed': 0,
        'unbooked': 0,
        'weeklyTarget': 0,
        'thisWeekUrgent': 0,
        'atRisk': 0,
        'byUrgency': empty_urgency_buckets(),
    }

def empty_portfolio_stats():
    return {
        'totalT1T2': 0,
        'completed': 0,
        'booked': 0,
        'outreachSent': 0,
        'overdue': 0,
        'upsellCandidates': 0,
        'byTier': {
            1: empty_tier_stats(),
            2: empty_tier_stats(),
            3: empty_tier_stats(),
        }
    }

def make_district(row, csm_slug, csm_name, tier_num, meeting_type):
    # Columns (post-shifts): A=name, B=training tier, C=live meeting Y/N,
    # D=account owner, E=active renewal, F=mpoc, G=LDoS, H=call target,
    # I=booked, J=meeting date, K=outreach sent,
    # L=async form sent on date (tier 3), M=last outreach, N=completed,
    # ..., U=notes
    name = row[0].strip()
    owner = row[3].strip()
    ldos = parse_date(row[6])
    bt = booking_target(ldos)
    booked_raw = row[8].strip()
    is_async_text = 'async' in booked_raw.lower() or 'async' in row[9].lower()
    booked = yn(booked_raw) and not is_async_text
    outreach = yn(row[10])
    async_form_sent = parse_date(row[11])
    last_outreach = parse_date(row[12])
    completed = yn(row[13])

    overdue = is_overdue(booked, completed, bt)
    nudge = needs_nudge(booked, completed, outreach, last_outreach)
    status = get_status(completed, meeting_type, overdue, booked, nudge)

    m = get_static_metrics(name)
    upsell = get_static_upsell(name)

    return {
        'name': name,
        'shortName': name.replace(' Unified School District', '').replace(' School District', '').replace(' Independent School District', '').strip(),
        'owner': owner,
        'csm': csm_slug,
        'csmName': csm_name,
        'tier': f'Tier {tier_num}',
        'tierNum': tier_num,
        'meetingType': meeting_type,
        'activeRenewal': yn(row[4]),
        'lastDayOfSchool': ldos.isoformat() if ldos else None,
        'bookingTarget': bt.isoformat() if bt else None,
        'booked': booked,
        'meetingDate': parse_date(row[9]).isoformat() if parse_date(row[9]) else None,
        'outreachSent': outreach,
        'asyncFormSent': async_form_sent.isoformat() if async_form_sent else None,
        'completed': completed,
        'notes': row[20] if len(row) > 20 else '',
        'status': status,
        'overdue': overdue,
        'needsNudge': nudge,
        'isUpsellCandidate': upsell is not None,
        'utilization': None,
        'upsellData': upsell,
        'mpocs': [],
        'enrollment': m.get('enrollment'),
        'ytdPacing': m.get('ytdPacing'),
    }

# Build districts
districts = []
orphans = []

for row in data_rows:
    while len(row) < 21:
        row.append('')
    name = row[0].strip()
    if not name:
        continue
    tier_num = parse_tier(row[1])
    meeting_type = parse_meeting_type(row[2], tier_num)
    owner = row[3].strip()
    csm_slug = csm_map.get(owner)
    if not csm_slug:
        orphans.append(make_district(row, 'unassigned', owner, tier_num, meeting_type))
        continue
    districts.append(make_district(row, csm_slug, owner, tier_num, meeting_type))

def build_urgency_buckets(unbooked_districts):
    buckets = empty_urgency_buckets()
    for d in unbooked_districts:
        bt = d.get('bookingTarget')
        if not bt:
            buckets['noDate'].append(d)
            continue
        days = (date.fromisoformat(bt) - today).days
        if days <= 0:
            buckets['thisWeek'].append(d)
        elif days <= 7:
            buckets['thisWeek'].append(d)
        elif days <= 14:
            buckets['nextWeek'].append(d)
        elif days <= 21:
            buckets['soon'].append(d)
        else:
            buckets['later'].append(d)
    return buckets

def build_gap_to_goal(t1t2_districts, all_districts):
    needs_call = [d for d in t1t2_districts]
    booked = [d for d in needs_call if d['booked']]
    completed = [d for d in all_districts if d['completed']]
    unbooked = [d for d in needs_call if not d['booked'] and not d['completed']]
    this_week_urgent = [d for d in unbooked if d.get('bookingTarget') and (date.fromisoformat(d['bookingTarget']) - today).days <= 7]
    at_risk = [d for d in t1t2_districts if d['overdue']]
    weekly_target = max(1, len(unbooked) // 3) if unbooked else 0

    return {
        'totalNeedingCall': len(needs_call),
        'booked': len(booked),
        'completed': len(completed),
        'unbooked': len(unbooked),
        'weeklyTarget': weekly_target,
        'thisWeekUrgent': len(this_week_urgent),
        'atRisk': len(at_risk),
        'byUrgency': build_urgency_buckets(unbooked),
    }

def build_portfolio_stats(all_districts, t1t2_districts):
    by_tier = {
        1: empty_tier_stats(),
        2: empty_tier_stats(),
        3: empty_tier_stats(),
    }
    for d in all_districts:
        tn = d['tierNum']
        by_tier[tn]['total'] += 1
        if d['completed']:
            by_tier[tn]['completed'] += 1
        elif d['booked']:
            by_tier[tn]['booked'] += 1
        else:
            by_tier[tn]['remaining'] += 1
        if d['overdue']:
            by_tier[tn]['overdue'] += 1

    return {
        'totalT1T2': len(t1t2_districts),
        'completed': len([d for d in t1t2_districts if d['completed']]),
        'booked': len([d for d in t1t2_districts if d['booked']]),
        'outreachSent': len([d for d in t1t2_districts if d['outreachSent']]),
        'overdue': len([d for d in t1t2_districts if d['overdue']]),
        'upsellCandidates': len([d for d in all_districts if d['isUpsellCandidate']]),
        'byTier': by_tier,
    }

# Build CSM data
csm_slugs = ['brianna', 'sarah', 'monica', 'daisy']
csm_data = {}

for slug in csm_slugs:
    csm_districts = [d for d in districts if d['csm'] == slug]
    t1t2 = [d for d in csm_districts if d['tierNum'] <= 2]
    csm_data[slug] = {
        'districts': csm_districts,
        'gapToGoal': build_gap_to_goal(t1t2, csm_districts),
        'stats': build_portfolio_stats(csm_districts, t1t2),
    }

# Portfolio
all_t1t2 = [d for d in districts if d['tierNum'] <= 2]

snapshot = {
    'refreshedAt': datetime.utcnow().isoformat() + 'Z',
    'stale': False,
    'portfolio': {
        'gapToGoal': build_gap_to_goal(all_t1t2, districts),
        'stats': build_portfolio_stats(districts, all_t1t2),
    },
    'csms': csm_data,
    'orphans': orphans,
}

os.makedirs('public/data', exist_ok=True)
with open('public/data/snapshot.json', 'w') as f:
    json.dump(snapshot, f, indent=2, default=str)

print(f"Done. {len(districts)} districts, {len(orphans)} orphans")
for slug in csm_slugs:
    s = csm_data[slug]['stats']
    print(f"  {slug}: {s['totalT1T2']} T1/T2, {s['booked']} booked, {s['overdue']} overdue")
