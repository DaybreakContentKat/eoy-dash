"""Generate public/data/bts.json — BTS Readiness data for the /bts page.

Runs in the nightly GitHub Action AFTER generate_snapshot.py. It:
  - reads public/data/snapshot.json for the district -> owner/tier/LDoS lookup
    (the tracker is already parsed there; we don't re-parse it here)
  - reads the BTS form-response sheet (two tabs) via the same service account
  - matches form districts to tracker districts (crosswalk + fuzzy suffix strip)
  - detects required fields per tab from the header (any "(required)" column)
  - computes per-district missing-field gaps + scheduling status
  - writes public/data/bts.json for app/bts/page.tsx to render at runtime

Pure logic (matching, parsing, building) is factored out of the Sheets read so
it can be dry-run locally without service-account credentials.
"""

import io
import json
import os
import re
import sys

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SNAPSHOT_PATH = os.path.join(REPO_ROOT, 'public', 'data', 'snapshot.json')
CROSSWALK_PATH = os.path.join(REPO_ROOT, 'district_crosswalk.json')
OUTPUT_PATH = os.path.join(REPO_ROOT, 'public', 'data', 'bts.json')
UNMATCHED_PATH = os.path.join(REPO_ROOT, 'unmatched_districts.txt')

BTS_FORM_SHEET_ID = '1ggBCh3KFvVIbz6EbzT9g1fczCsyoXx5cKYciNzSvyPs'
CSM_TAB = 'CSM / AM filled out'   # T1/T2 live-meeting districts
ASYNC_TAB = 'Async Version'       # T3 districts

CSM_FULLNAME = {
    'brianna': 'Brianna Masciel',
    'sarah': 'Sarah Hough',
    'monica': 'Monica Knott',
    'daisy': 'Daisy Leahy',
}
OWNER_ORDER = ['Brianna Masciel', 'Sarah Hough', 'Monica Knott', 'Daisy Leahy']

# Normalize the self-reported Account Owner field on form responses. Split
# entries ("daisy/sarah") resolve to the first listed name and flag co_owned.
OWNER_NORMALIZE = {
    'bri': 'Brianna Masciel',
    'brianna': 'Brianna Masciel',
    'brianna masciel': 'Brianna Masciel',
    'sarah': 'Sarah Hough',
    'sarah h': 'Sarah Hough',
    'sarah hough': 'Sarah Hough',
    'monica': 'Monica Knott',
    'monica knott': 'Monica Knott',
    'daisy': 'Daisy Leahy',
    'daisy leahy': 'Daisy Leahy',
}

# Full column header -> short label for gap display.
SHORT_LABELS = [
    ('summer comms discussed', 'Summer comms discussed'),
    ('summer resource post', 'Summer resource post'),
    ('staff training date', 'Training date'),
    ('staff training scheduled', 'Training scheduled'),
    ('admin / principal outreach', 'Admin/principal outreach'),
    ('principal contact list', 'Principal contact list'),
    ('staff files status', 'Staff files status'),
    ('are you planning to send a message to families', 'Family summer comms plan'),
    ('who is the best person on your team to reach', 'Summer contact'),
    ('each year we update the list of staff', 'Staff list update'),
]

# Scheduling fields (CSM tab only), matched by a substring of the header.
SCHED_KICKOFF = 'kickoff with leadership'
SCHED_TRAINING_DATE = 'staff training date'
SCHED_TRAINING_SCHEDULED = 'staff training scheduled'
SCHED_STAFF_FILE = 'staff files status'
SCHED_FAMILY_COMMS = 'are you planning to send a message to families'  # async "complete" preview

# Free-text columns fed to the Claude synthesis (lib/prompts side). Each key maps
# to the substring needles to try in order; the CSM/AM and Async tabs word these
# questions differently, so several keys list a per-tab fallback.
SYNTH_NEEDLES = {
    'teacher_resources': ('teacher resource expression of interest',),  # CSM/AM tab only
    'outstanding_items': ('outstanding items still needed',),           # CSM/AM tab only
    'feedback': ('feedback',),  # CSM "Feedback"; Async "...feedback on last year?"
    'comms_channels': ('district comms channels',
                       'how does your district usually communicate'),
    'family_comms_plan': ('are you planning to send a message to families',),  # Async tab
}
SYNTH_KEYS = tuple(SYNTH_NEEDLES)

SUFFIXES = [
    'independent school district', 'unified school district', 'elementary school district',
    'union high school district', 'community unit school district', 'joint unified school district',
    'county office of education', 'public schools', 'county schools', 'school district',
    'school system', 'college preparatory', 'schools', 'district', 'isd', 'usd', 'cusd',
]


# --------------------------------------------------------------------------
# Pure helpers (no Sheets / no I/O beyond the file reads passed in)
# --------------------------------------------------------------------------

def norm(s):
    s = re.sub(r'[.,()]', ' ', (s or '').strip().lower())
    return re.sub(r'\s+', ' ', s).strip()


def strip_suffix(s):
    s = norm(s)
    changed = True
    while changed:
        changed = False
        for suf in SUFFIXES:
            if s.endswith(' ' + suf) or s == suf:
                s = s[: -len(suf)].strip()
                changed = True
    return s.strip()


def short_label(header):
    h = norm(header)
    for needle, label in SHORT_LABELS:
        if needle in h:
            return label
    return re.sub(r'\s*\(required\)\s*', '', header, flags=re.I).strip()


def status_state(value):
    """Map a scheduling cell to a display state. Confirmed=green,
    Tentative=amber, Not Yet/blank=red, any other text=amber (in-progress)."""
    v = (value or '').strip()
    if not v:
        return 'red'
    low = v.lower()
    if low == 'confirmed' or low.startswith('confirmed'):
        return 'green'
    if 'tentative' in low:
        return 'amber'
    if low.startswith('not yet') or low == 'no' or low == 'n/a':
        return 'red'
    return 'amber'


# A required field counts as a GAP unless it has a firm, affirmative answer.
# Blanks, placeholders ("TBD"/"Pending"/"Discuss later"), and in-progress
# statuses ("Not Yet"/"Tentative Hold"/"Partial"/"New File coming") all flag.
# The actual entered value is always shown next to the flag, so the human can
# eyeball any mis-flag — these heuristics just decide the amber highlight.
COMPLETE_PREFIXES = ('yes', 'confirmed', 'received', 'complete', 'done', 'already')
NOTDONE_PATTERNS = [
    'tbd', 'pending', 'n/a', 'not yet', 'tentative', 'partial', 'new file coming',
    'file coming', 'to discuss', 'discuss in', 'discuss after', "let's discuss",
    'lets discuss', 'will gather', 'will discuss', 'will set', 'will send',
    'will confirm', 'will look', 'will get', 'will check', 'will let', 'will have',
    'will ask', 'will plan', 'soon', 'unknown', 'maybe', 'not sure', 'need to discuss',
    'not needed', 'not right now', 'kind of', '?', 'none', 'na',
]
_MONTHS = (r'jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec')


def _has_real_date(v):
    """True if the value names a concrete day (7/23, 8-11, 'August 7th', '16th'),
    not a vague window ('Late August', 'End of July', 'July', 'TBD')."""
    if re.search(r'\b\d{1,2}\s*[/-]\s*\d{1,2}', v):
        return True
    if re.search(r'(' + _MONTHS + r')[a-z]*\.?\s+\d{1,2}', v, re.I):
        return True
    if re.search(r'\b\d{1,2}(st|nd|rd|th)\b', v):
        return True
    return False


def field_gap(header, value):
    """Return True if this required-field value is a gap (no firm answer)."""
    h = norm(header)
    v = (value or '').strip()
    low = v.lower()
    if not v:
        return True
    if 'staff training date' in h:
        return not _has_real_date(v)
    if low.startswith(COMPLETE_PREFIXES) or 'use this year' in low:
        return False
    if any(p in low for p in NOTDONE_PATTERNS):
        return True
    if low in ('no', 'na', 'n/a', 'none'):
        return True
    return False


def normalize_owner(raw):
    """Return (canonical_full_name_or_None, co_owned_bool)."""
    if not raw:
        return None, False
    parts = re.split(r'[/&]| and ', raw)
    co_owned = len([p for p in parts if p.strip()]) > 1
    first = norm(parts[0])
    return OWNER_NORMALIZE.get(first), co_owned


def load_tracker(snapshot):
    """district canonical name -> {owner, tier, ldos, shortName, coOwned}.
    Universe = the 4 CSMs' portfolios (churned/onsite orphans excluded)."""
    tracker = {}
    for slug, csm in snapshot.get('csms', {}).items():
        owner = CSM_FULLNAME.get(slug, slug)
        for d in csm.get('districts', []):
            tracker[d['name']] = {
                'owner': owner,
                'tier': d['tierNum'],
                'ldos': d.get('lastDayOfSchool'),
                'shortName': d.get('shortName') or d['name'],
                'coOwned': '/' in (d.get('owner') or ''),
            }
    return tracker


def build_matcher(tracker, crosswalk):
    uni_norm = {norm(k): k for k in tracker}
    uni_strip = {}
    for k in tracker:
        uni_strip.setdefault(strip_suffix(k), k)

    def match(name):
        n = norm(name)
        if n in crosswalk:                       # Layer 1: crosswalk exact
            cand = crosswalk[n]
            return (cand if cand in tracker else cand), 'crosswalk'
        if n in uni_norm:
            return uni_norm[n], 'exact'
        s = strip_suffix(name)                   # Layer 2: fuzzy suffix strip
        if s in uni_strip:
            return uni_strip[s], 'fuzzy'
        cands = [orig for st, orig in uni_strip.items() if st and (st in s or s in st)]
        if len(cands) == 1:
            return cands[0], 'fuzzy'
        return None, 'unmatched'

    return match


def parse_tab(header, rows, is_csm_tab):
    """Return (required_cols, parsed_responses) for one tab.

    header: list of column names. rows: list of cell-value lists.
    Each response: dict with raw_name, owner_raw, missing_fields (short labels),
    gap_count, scheduling fields, family_comms, timestamp, is_csm_tab.
    """
    idx = {norm(h): i for i, h in enumerate(header)}

    def col(needle):
        for h, i in idx.items():
            if needle in h:
                return i
        return None

    district_i = col('district name')
    owner_i = col('account owner')
    ts_i = col('timestamp')
    required = [(i, h) for i, h in enumerate(header) if '(required)' in h.lower()]

    sched_i = {
        'kickoff': col(SCHED_KICKOFF),
        'training_date': col(SCHED_TRAINING_DATE),
        'training_scheduled': col(SCHED_TRAINING_SCHEDULED),
        'staff_file': col(SCHED_STAFF_FILE),
        'family_comms': col(SCHED_FAMILY_COMMS),
    }

    def col_any(*needles):
        for needle in needles:
            i = col(needle)
            if i is not None:
                return i
        return None

    # Free-text fields fed to the Claude synthesis. The two tabs phrase these
    # differently, so each key lists the needles to try in order (first hit wins).
    synth_i = {k: col_any(*needles) for k, needles in SYNTH_NEEDLES.items()}

    def cell(row, i):
        if i is None or i >= len(row):
            return ''
        return (row[i] or '').strip()

    responses = []
    for row in rows:
        name = cell(row, district_i)
        if not name:
            continue  # skip blank district rows silently
        gaps = [{'field': short_label(h), 'value': cell(row, i)}
                for i, h in required if field_gap(h, cell(row, i))]
        resp = {
            'raw_name': name,
            'owner_raw': cell(row, owner_i),
            'timestamp': cell(row, ts_i),
            'gaps': gaps,
            'missing_fields': [g['field'] for g in gaps],  # back-compat
            'gap_count': len(gaps),
            'is_csm_tab': is_csm_tab,
            'synth': {k: cell(row, synth_i[k]) for k in SYNTH_KEYS},
        }
        if is_csm_tab:
            resp['training_date'] = cell(row, sched_i['training_date'])
            resp['training_scheduled'] = cell(row, sched_i['training_scheduled'])
            resp['kickoff_status'] = cell(row, sched_i['kickoff'])
            resp['staff_file_status'] = cell(row, sched_i['staff_file'])
        else:
            resp['family_comms'] = cell(row, sched_i['family_comms'])
        responses.append(resp)
    return [h for _, h in required], responses


def _ts_key(resp):
    # Form timestamps look like "5/14/2026 13:16:09"; sort lexicographically
    # after zero-padding is overkill — parse to a sortable tuple, fall back to 0.
    m = re.match(r'(\d+)/(\d+)/(\d+)\s+(\d+):(\d+):(\d+)', resp.get('timestamp', '') or '')
    if not m:
        return (0,) * 6
    mo, da, yr, hh, mm, ss = (int(x) for x in m.groups())
    return (yr, mo, da, hh, mm, ss)


def choose_response(responses, tier):
    """Among responses for one district, prefer the tab matching its tier
    (T1/T2 -> CSM tab, T3 -> async tab), then latest timestamp."""
    if not responses:
        return None
    prefer_csm = tier in (1, 2)
    preferred = [r for r in responses if r['is_csm_tab'] == prefer_csm]
    pool = preferred or responses
    return sorted(pool, key=_ts_key)[-1]


def build_bts(tracker, match, csm_responses, async_responses, refreshed_at):
    all_responses = csm_responses + async_responses
    # group matched responses by canonical district; collect unmatched
    by_district = {}
    unmatched = []
    for r in all_responses:
        canon, how = match(r['raw_name'])
        if canon and canon in tracker:
            by_district.setdefault(canon, []).append(r)
        else:
            unmatched.append(r)

    def new_owner():
        return {
            'noForm': [], 'missing': [], 'complete': [], 'submitted': [],
            # per-tier summary: districts owned, no-form, submitted, submitted-with-gaps
            'byTier': {str(t): {'total': 0, 'noForm': 0, 'submitted': 0, 'withGaps': 0}
                       for t in (1, 2, 3)},
        }

    owners = {name: new_owner() for name in OWNER_ORDER}
    owners['Unknown'] = new_owner()

    def sched(resp, key):
        val = resp.get(key, '') if resp else ''
        return {'label': val or '—', 'state': status_state(val)}

    for canon, meta in tracker.items():
        owner = meta['owner'] if meta['owner'] in owners else 'Unknown'
        tier = meta['tier']
        bt = owners[owner]['byTier'].get(str(tier))
        if bt is not None:
            bt['total'] += 1
        resp = choose_response(by_district.get(canon, []), tier)
        if resp is None:
            owners[owner]['noForm'].append({
                'name': canon, 'shortName': meta['shortName'],
                'tier': tier, 'ldos': meta['ldos'],
            })
            if bt is not None:
                bt['noForm'] += 1
            continue
        form_owner, form_co = normalize_owner(resp.get('owner_raw'))
        co_owned = bool(meta['coOwned'] or form_co)
        # Surface when the form's self-reported owner disagrees with the tracker
        # owner (the authoritative one we group by) — flags tracker data to check.
        raw_fo = (resp.get('owner_raw') or '').strip()
        display_fo = form_owner or raw_fo
        form_owner_note = display_fo if (display_fo and norm(display_fo) != norm(meta['owner'])) else None
        has_gaps = resp['gap_count'] > 0

        # A submitted form always counts as submitted; gaps are an overlay flag.
        sub = {
            'name': canon, 'tier': tier, 'gapCount': resp['gap_count'],
            'gaps': resp.get('gaps', []),  # [{field, value}] — what's missing + what they put
            'coOwned': co_owned, 'formOwner': form_owner_note,
        }
        if resp['is_csm_tab']:
            sub['trainingStatus'] = sched(resp, 'training_scheduled')
            sub['kickoffStatus'] = sched(resp, 'kickoff_status')
            sub['staffFileStatus'] = sched(resp, 'staff_file_status')
            sub['trainingDate'] = resp.get('training_date', '') or ''
        else:
            fc = (resp.get('family_comms') or '').strip()
            sub['familyComms'] = (fc[:40] + '…') if len(fc) > 40 else fc
        owners[owner]['submitted'].append(sub)
        if bt is not None:
            bt['submitted'] += 1
            if has_gaps:
                bt['withGaps'] += 1

        # --- back-compat: keep populating missing/complete (old frontend reads these) ---
        if has_gaps:
            owners[owner]['missing'].append({
                'name': canon, 'tier': tier, 'gapCount': resp['gap_count'],
                'missingFields': resp['missing_fields'], 'gaps': resp.get('gaps', []),
                'trainingStatus': sched(resp, 'training_scheduled') if resp['is_csm_tab'] else None,
                'kickoffStatus': sched(resp, 'kickoff_status') if resp['is_csm_tab'] else None,
                'coOwned': co_owned, 'formOwner': form_owner_note,
            })
        else:
            entry = {'name': canon, 'tier': tier, 'coOwned': co_owned, 'formOwner': form_owner_note}
            if resp['is_csm_tab']:
                entry['trainingStatus'] = sched(resp, 'training_scheduled')
                entry['kickoffStatus'] = sched(resp, 'kickoff_status')
                entry['staffFileStatus'] = sched(resp, 'staff_file_status')
            else:
                fc = (resp.get('family_comms') or '').strip()
                entry['familyComms'] = (fc[:40] + '…') if len(fc) > 40 else fc
            owners[owner]['complete'].append(entry)

    # unmatched form districts -> Unknown owner, yellow flag
    for r in unmatched:
        owners['Unknown']['submitted'].append({
            'name': r['raw_name'], 'tier': 'Unknown', 'gapCount': r['gap_count'],
            'gaps': r.get('gaps', []), 'coOwned': False, 'unmatched': True,
        })
        owners['Unknown']['missing'].append({
            'name': r['raw_name'], 'tier': 'Unknown', 'gapCount': r['gap_count'],
            'missingFields': r['missing_fields'], 'gaps': r.get('gaps', []),
            'trainingStatus': None, 'kickoffStatus': None,
            'coOwned': False, 'unmatched': True,
        })

    # sort groups
    for o in owners.values():
        o['noForm'].sort(key=lambda d: (d['ldos'] is None, d['ldos'] or ''))
        o['missing'].sort(key=lambda d: -d['gapCount'])
        o['complete'].sort(key=lambda d: d['name'])
        # submitted: worst gaps first, then by tier, then name
        o['submitted'].sort(key=lambda d: (-d['gapCount'], d['tier'] if isinstance(d['tier'], int) else 9, d['name']))

    # scheduling summary table (T1/T2 districts with a CSM-tab response)
    scheduling = []
    for canon, meta in tracker.items():
        if meta['tier'] not in (1, 2):
            continue
        resp = choose_response(by_district.get(canon, []), meta['tier'])
        if not resp or not resp['is_csm_tab']:
            continue
        scheduling.append({
            'district': canon, 'owner': meta['owner'],
            'trainingScheduled': sched(resp, 'training_scheduled'),
            'trainingDate': resp.get('training_date', '') or '—',
            'kickoffStatus': sched(resp, 'kickoff_status'),
            'staffFileStatus': sched(resp, 'staff_file_status'),
        })
    scheduling.sort(key=lambda r: (r['owner'], r['district']))

    # totals / stat pills
    universe = list(tracker.values())
    t1t2 = [m for m in universe if m['tier'] in (1, 2)]
    t3 = [m for m in universe if m['tier'] == 3]
    matched_canon = set(by_district)

    def submitted(metas):
        return sum(1 for canon, m in tracker.items() if m in metas and canon in matched_canon)

    # complete = response present with zero gaps
    complete_canon = {
        canon for canon, resps in by_district.items()
        if (choose_response(resps, tracker[canon]['tier']) or {}).get('gap_count', 1) == 0
    }
    with_gaps = sum(
        1 for canon in matched_canon
        if (choose_response(by_district[canon], tracker[canon]['tier']) or {}).get('gap_count', 0) > 0
    )
    # portfolio-wide tier summary (sum of per-owner byTier across the 4 CSMs)
    tier_summary = {str(t): {'total': 0, 'noForm': 0, 'submitted': 0, 'withGaps': 0}
                    for t in (1, 2, 3)}
    for name in OWNER_ORDER:
        for t, c in owners[name]['byTier'].items():
            for k in ('total', 'noForm', 'submitted', 'withGaps'):
                tier_summary[t][k] += c[k]

    totals = {
        'totalDistricts': len(universe),
        'formsSubmitted': len(matched_canon),
        'formsWithGaps': with_gaps,
        'formsClean': len(matched_canon) - with_gaps,
        't1t2Total': len(t1t2),
        't1t2Complete': sum(1 for c in complete_canon if tracker[c]['tier'] in (1, 2)),
        't3Total': len(t3),
        't3Complete': sum(1 for c in complete_canon if tracker[c]['tier'] == 3),
        'withGaps': with_gaps,
        'unmatchedCount': len(unmatched),
        'tierSummary': tier_summary,
    }

    return {
        'refreshedAt': refreshed_at,
        'totals': totals,
        'ownerOrder': OWNER_ORDER + (['Unknown'] if (owners['Unknown']['submitted'] or
                                                     owners['Unknown']['noForm']) else []),
        'owners': owners,
        'scheduling': scheduling,
        'unmatched': sorted({r['raw_name'] for r in unmatched}),
    }


# --------------------------------------------------------------------------
# Claude synthesis — turn the free-text form fields into CSM-facing themes
# --------------------------------------------------------------------------

SYNTH_SYSTEM = (
    "You are summarizing BTS planning form responses for a school mental "
    "health company's customer success team. Output valid JSON only. "
    "No preamble, no markdown fences, no explanation outside the JSON."
)


def build_synthesis_rows(all_responses, match, tracker):
    """One dict per form row that has at least one non-blank synthesis field.
    District/owner resolve through the tracker when matched, else fall back to
    the form's own self-reported values."""
    rows = []
    for r in all_responses:
        synth = r.get('synth', {})
        if not any((synth.get(k) or '').strip() for k in SYNTH_KEYS):
            continue
        canon, _ = match(r['raw_name'])
        if canon and canon in tracker:
            district = tracker[canon]['shortName'] or canon
            owner = tracker[canon]['owner']
        else:
            district = r['raw_name']
            form_owner, _ = normalize_owner(r.get('owner_raw'))
            owner = form_owner or (r.get('owner_raw') or '').strip() or 'Unknown'
        row = {'district': district, 'owner': owner}
        row.update({k: (synth.get(k) or '').strip() for k in SYNTH_KEYS})
        rows.append(row)
    return rows


def synthesize(form_rows_for_synthesis):
    """Call Claude to theme the free-text responses. Returns the parsed JSON
    dict, or None on any failure (missing key, network, bad JSON) — the caller
    treats None as "no synthesis this run" and the pipeline continues."""
    import urllib.error
    import urllib.request

    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        print('WARNING: ANTHROPIC_API_KEY not set, skipping BTS synthesis',
              file=sys.stderr)
        return None
    if not form_rows_for_synthesis:
        print('WARNING: no synthesis rows to send, skipping BTS synthesis',
              file=sys.stderr)
        return None

    user_prompt = (
        "Analyze these BTS planning form responses and return a JSON object "
        "with exactly these four keys:\n\n"
        "teacher_themes: Array of up to 5 objects:\n"
        "  { label, explanation, districts[] }\n"
        "  Derive from the Teacher Resource Expression of Interest field.\n"
        "  Skip blank, N/A, No responses.\n\n"
        "comms_channels: Array of objects sorted by count descending:\n"
        "  { channel, count, districts[] }\n"
        "  Normalize similar labels (e.g. ParentSquare/Peachjar/text platform).\n\n"
        "outstanding_items: Array of objects, one per district with substantive "
        "content in the Outstanding items still needed field. Skip NA/None/blank:\n"
        "  { district, owner, items[] }\n"
        "  items[] = short action strings, max 12 words each.\n\n"
        "feedback_themes: Array of up to 4 objects:\n"
        "  { theme, sentiment (working_well | needs_improvement | mixed), examples[] }\n"
        "  examples[] = up to 3 paraphrases max 20 words each with district attribution.\n\n"
        "Form responses: " + json.dumps(form_rows_for_synthesis)
    )

    body = json.dumps({
        'model': 'claude-sonnet-4-6',
        'max_tokens': 2000,
        'system': SYNTH_SYSTEM,
        'messages': [{'role': 'user', 'content': user_prompt}],
    }).encode('utf-8')

    req = urllib.request.Request(
        'https://api.anthropic.com/v1/messages',
        data=body,
        headers={
            'x-api-key': api_key,
            'anthropic-version': '2023-06-01',
            'content-type': 'application/json',
        },
        method='POST',
    )

    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            payload = json.loads(resp.read().decode('utf-8'))
        text = ''.join(
            block.get('text', '')
            for block in payload.get('content', [])
            if block.get('type') == 'text'
        )
        result = json.loads(text)
        print(f'BTS synthesis: ok ({len(form_rows_for_synthesis)} rows -> '
              f'{", ".join(result.keys())})')
        return result
    except urllib.error.HTTPError as e:
        detail = e.read().decode('utf-8', 'replace')[:500]
        print(f'WARNING: BTS synthesis HTTP {e.code}: {detail}', file=sys.stderr)
        return None
    except Exception as e:
        print(f'WARNING: BTS synthesis failed ({type(e).__name__}): {e}',
              file=sys.stderr)
        return None


# --------------------------------------------------------------------------
# Sheets read (the only part that needs service-account credentials)
# --------------------------------------------------------------------------

def read_form_tabs():
    """Return {tab_name: (header, rows)} for the BTS form sheet."""
    from google.oauth2.service_account import Credentials
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaIoBaseDownload
    import openpyxl

    creds = Credentials.from_service_account_info(
        json.loads(os.environ['GOOGLE_SERVICE_ACCOUNT_KEY']),
        scopes=['https://www.googleapis.com/auth/drive.readonly'],
    )
    drive = build('drive', 'v3', credentials=creds)
    xlsx = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    meta = drive.files().get(fileId=BTS_FORM_SHEET_ID, fields='mimeType').execute()
    if meta['mimeType'] == 'application/vnd.google-apps.spreadsheet':
        request = drive.files().export_media(fileId=BTS_FORM_SHEET_ID, mimeType=xlsx)
    else:
        request = drive.files().get_media(fileId=BTS_FORM_SHEET_ID)
    buf = io.BytesIO()
    dl = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = dl.next_chunk()
    buf.seek(0)
    wb = openpyxl.load_workbook(buf, data_only=True)
    out = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows = [[('' if c is None else str(c)) for c in r] for r in ws.iter_rows(values_only=True)]
        if rows:
            out[name] = (rows[0], rows[1:])
    return out


def find_tab(tabs, wanted):
    for name, data in tabs.items():
        if norm(name) == norm(wanted):
            return data
    print(f'WARNING: tab not found by name: {wanted!r}', file=sys.stderr)
    return None


def classify_tab(header):
    """Identify a form tab by its header signature rather than its sheet name.

    The sheet is a Google Form responses workbook; tab names drift (renames,
    "Form Responses N" defaults), so matching on the columns each form owns is
    far more robust. Returns 'csm', 'async', or None.
      - CSM/AM tab: has an "Account Owner" column AND "Summer comms discussed?".
      - Async tab:  has the "...send a message to families..." column.
    """
    hs = [norm(h) for h in header]
    has = lambda needle: any(needle in h for h in hs)
    if has('account owner') and has('summer comms discussed'):
        return 'csm'
    if has('are you planning to send a message to families'):
        return 'async'
    return None


def select_tabs(tabs):
    """Return (csm_tab_data, async_tab_data), each (header, rows) or None.
    Prefer header-signature detection; fall back to exact-name lookup."""
    csm_tab = async_tab = None
    for name, data in tabs.items():
        kind = classify_tab(data[0])
        if kind == 'csm' and csm_tab is None:
            csm_tab = data
            print(f'BTS: CSM/AM tab matched by header signature: {name!r}')
        elif kind == 'async' and async_tab is None:
            async_tab = data
            print(f'BTS: Async tab matched by header signature: {name!r}')
    if csm_tab is None:
        csm_tab = find_tab(tabs, CSM_TAB)
    if async_tab is None:
        async_tab = find_tab(tabs, ASYNC_TAB)
    if csm_tab is None:
        print('ERROR: could not locate the CSM/AM form tab (by signature or name)',
              file=sys.stderr)
    if async_tab is None:
        print('ERROR: could not locate the Async form tab (by signature or name)',
              file=sys.stderr)
    return csm_tab, async_tab


def load_dotenv_local():
    """Local-dev convenience: pull ANTHROPIC_API_KEY (and friends) from
    scripts/.env.local if not already in the environment. In CI the value comes
    from the workflow's env block, so this is a no-op there."""
    env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '.env.local')
    if not os.path.exists(env_path):
        return
    with open(env_path) as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith('#') or '=' not in line:
                continue
            key, _, val = line.partition('=')
            os.environ.setdefault(key.strip(), val.strip().strip('"').strip("'"))


def main():
    load_dotenv_local()
    with open(SNAPSHOT_PATH) as f:
        snapshot = json.load(f)
    with open(CROSSWALK_PATH) as f:
        crosswalk = json.load(f)
    tracker = load_tracker(snapshot)
    match = build_matcher(tracker, crosswalk)

    try:
        tabs = read_form_tabs()
    except Exception as e:  # Sheets failure must not crash the nightly routine
        print(f'ERROR: could not read BTS form sheet, skipping BTS generation: {e}',
              file=sys.stderr)
        return 0

    csm_tab, async_tab = select_tabs(tabs)
    csm_responses = parse_tab(csm_tab[0], csm_tab[1], True)[1] if csm_tab else []
    async_responses = parse_tab(async_tab[0], async_tab[1], False)[1] if async_tab else []

    bts = build_bts(tracker, match, csm_responses, async_responses,
                    snapshot.get('refreshedAt', ''))

    # Claude synthesis of the free-text fields. Failure here is non-fatal:
    # synthesize() returns None and we write bts.json without it.
    synth_rows = build_synthesis_rows(csm_responses + async_responses, match, tracker)
    bts['synthesis'] = synthesize(synth_rows)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, 'w') as f:
        json.dump(bts, f, indent=2)

    if bts['unmatched']:
        with open(UNMATCHED_PATH, 'w') as f:
            f.write('\n'.join(bts['unmatched']) + '\n')
        print(f'WARNING: {len(bts["unmatched"])} unmatched district(s): '
              f'{", ".join(bts["unmatched"])}', file=sys.stderr)

    t = bts['totals']
    print(f'Done. {t["formsSubmitted"]}/{t["totalDistricts"]} forms, '
          f'{t["withGaps"]} with gaps, {t["unmatchedCount"]} unmatched.')
    return 0


if __name__ == '__main__':
    sys.exit(main())
